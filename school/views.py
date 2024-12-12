from django.http import JsonResponse,HttpResponse
from rest_framework import status
# from weasyprint import HTML
from rest_framework.views import APIView

from student.models import Students,SchoolStudent


from payment.models import (ReceiptDetail,
                            student_fees,fee_type_master)


from .serializers import SchoolStudentSerializer,SchoolStudentDetailSerializer,SchoolStudentPostSerializer
from rest_framework.permissions import IsAuthenticated,BasePermission
from rest_framework_simplejwt.authentication import JWTAuthentication

from payment.views import HasFeeReportPermission
from django.views.generic import TemplateView
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from standard.models import AcademicYear

from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.models import Group, User
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
import uuid
from django.core.cache import cache
import datetime  # Add this import at the top




from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.db.models import Q
from .models import ChatMessage
from django.contrib.auth.models import User
from .serializers import ChatMessageSerializer, UserSerializer
from django.db.models import Max, Q
from django.db.models.functions import Coalesce
class ChatListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Annotate each user with the timestamp of the last message exchanged with the current user
        chat_users = User.objects.filter(
            is_active=True
        ).exclude(
            id=request.user.id  # Exclude current user
        ).annotate(
            last_sent_time=Max(
                'sent_messages__timestamp',
                filter=Q(sent_messages__receiver=request.user, sent_messages__deleted_by_sender=False)
            ),
            last_received_time=Max(
                'received_messages__timestamp',
                filter=Q(received_messages__sender=request.user, received_messages__deleted_by_receiver=False)
            )
        ).annotate(
            last_communication_time=Coalesce('last_sent_time', 'last_received_time')
        ).order_by('-last_communication_time')  # Order by the last communication time descending

        chat_data = []
        for user in chat_users:
            last_message = ChatMessage.objects.filter(
                (Q(sender=request.user, receiver=user, deleted_by_sender=False) |
                 Q(sender=user, receiver=request.user, deleted_by_receiver=False))
            ).last()
            
            chat_data.append({
                'user': UserSerializer(user).data,
                'last_message': ChatMessageSerializer(last_message).data if last_message else None,
                'unread_count': ChatMessage.objects.filter(
                    sender=user,
                    receiver=request.user,
                    is_read=False
                ).count()
            })
        
        return Response(chat_data)

class ChatMessageView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id):
        messages = ChatMessage.objects.filter(
            # Get messages where current user is sender and hasn't deleted them
            (Q(sender=request.user, receiver_id=user_id, deleted_by_sender=False) |
            # Or where current user is receiver and hasn't deleted them
            Q(sender_id=user_id, receiver=request.user, deleted_by_receiver=False))
            # Exclude messages deleted by both users
            & ~Q(deleted_by_sender=True, deleted_by_receiver=True)
        )
        serializer = ChatMessageSerializer(messages, many=True)
        return Response(serializer.data)

    def post(self, request, user_id):
        message = request.data.get('message')
        if not message:
            return Response({'error': 'Message is required'}, status=400)

        chat_message = ChatMessage.objects.create(
            sender=request.user,
            receiver_id=user_id,
            message=message
        )
        serializer = ChatMessageSerializer(chat_message)
        return Response(serializer.data, status=201)


class MarkChatAsReadView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, user_id):
        try:
            # Mark all messages from the specified user as read
            ChatMessage.objects.filter(
                sender_id=user_id,
                receiver=request.user,
                is_read=False
            ).update(is_read=True)
            
            return Response({'status': 'success'})
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class MessageStatusView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def patch(self, request, message_id, status_type=None):
        try:
            message = ChatMessage.objects.get(id=message_id)
            
            # Get status type from URL parameter
            status_type = self.kwargs.get('status_type')
            
            if status_type == 'delivered':
                message.is_delivered = True
            elif status_type == 'read':
                message.is_delivered = True
                message.is_read = True
            
            message.save()
            
            return Response({
                'id': message.id,
                'is_delivered': message.is_delivered,
                'is_read': message.is_read
            })
        except ChatMessage.DoesNotExist:
            return Response(
                {'error': 'Message not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MessageActionView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request, message_id):
        try:
            # Allow deletion if user is either sender or receiver
            message = ChatMessage.objects.get(
                Q(id=message_id),
                Q(sender=request.user) | Q(receiver=request.user)
            )
            message.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ChatMessage.DoesNotExist:
            return Response(
                {'error': 'Message not found or unauthorized'}, 
                status=status.HTTP_404_NOT_FOUND
            )

    def patch(self, request, message_id):
        try:
            message = ChatMessage.objects.get(
                id=message_id,
                sender=request.user  # Only allow editing own messages
            )
            new_message = request.data.get('message')
            if not new_message:
                return Response(
                    {'error': 'Message content is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            message.message = new_message
            message.edited = True  # Add this field to your model if not exists
            message.save()
            
            serializer = ChatMessageSerializer(message)
            return Response(serializer.data)
        except ChatMessage.DoesNotExist:
            return Response(
                {'error': 'Message not found or unauthorized'}, 
                status=status.HTTP_404_NOT_FOUND
            )
            
class BulkMessageDeleteView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        message_ids = request.data.get('message_ids', [])
        if not message_ids:
            return Response(
                {'error': 'No message IDs provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Allow deletion if user is either sender or receiver
        messages_to_delete = ChatMessage.objects.filter(
            Q(id__in=message_ids),
            Q(sender=request.user) | Q(receiver=request.user)
        )
        deleted_count, _ = messages_to_delete.delete()
        
        return Response({'deleted_count': deleted_count})



class ClearChatView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request, user_id):
        try:
            # Delete all messages between the two users
            deleted_count = ChatMessage.objects.filter(
                Q(sender=request.user, receiver_id=user_id) |
                Q(sender_id=user_id, receiver=request.user)
            ).delete()[0]
            
            return Response({
                'status': 'success',
                'deleted_count': deleted_count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
class HomePageView(TemplateView):
    template_name = "home.html"  


# permission for school student for Group
class HasSchoolStudentPermission(BasePermission):
    def has_permission(self, request, view):
        return request.user.has_perm(f'student.{view.required_permission}')

# get api for school student
class SchoolStudentGet(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasSchoolStudentPermission]
    required_permission = 'can_view_school_students'
    def get(self, request):
        
        school_student = SchoolStudent.objects.all()
        serialized_data = SchoolStudentSerializer(school_student, many=True)
        return JsonResponse({"message": "School Students Get Successfully", "data": serialized_data.data}, status=200)

# get api for student name

class SchoolStudentNamesGet(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasSchoolStudentPermission]
    required_permission = 'can_view_school_student_names'

    def get(self, request):
        
        school_student = Students.objects.all()
        serialized_data = SchoolStudentDetailSerializer(school_student, many=True)
        return JsonResponse({"message": "School Students Name Get Successfully", "data": serialized_data.data}, status=200)

# get by id api for school student 

class SchoolStudentByIdGet(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasSchoolStudentPermission]
    required_permission = 'can_view_school_student_details'


    def get(self, request, pk):
        try:
            school_student = SchoolStudent.objects.get(pk=pk)
        except SchoolStudent.DoesNotExist:
            return JsonResponse({"error": "School Student not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serialized_data = SchoolStudentSerializer(school_student)
        return JsonResponse({"message": "School Student Retrieved Successfully", "data": serialized_data.data}, status=status.HTTP_200_OK)

#  post api for school student

class SchoolStudentPost(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasSchoolStudentPermission]
    required_permission = 'can_add_school_student'


    def post(self, request):
        serializer = SchoolStudentPostSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse({"message": "School Student Created Successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# patch api for school student

class SchoolStudentPatch(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasSchoolStudentPermission]
    required_permission = 'can_edit_school_student'


    def patch(self, request, pk):
        try:
            school_student = SchoolStudent.objects.get(pk=pk)
        except SchoolStudent.DoesNotExist:
            return JsonResponse({"error": "School Student not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = SchoolStudentPostSerializer(school_student, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            return JsonResponse({"message": "School Student Updated Successfully", "data": serializer.data}, status=status.HTTP_200_OK)
        
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)






from payment.models import  student_fees,Receipt
class FeeReportDetailAPIViewDemo(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasFeeReportPermission]
    required_permission = 'can_view_fee_report_detail'

    def get(self, request, standard):
        try:
            # Get student fees data with fee type details
            student_fees_queryset = student_fees.objects.filter(
                standard__name=standard,
                is_assigned=True,
                student__academic_year=AcademicYear.objects.filter(is_current=True).first()
            ).select_related(
                'student', 
                'standard', 
                'fee_type', 
                'fee_type__fee_master'
            )

            # Get receipt details
            receipt_details = ReceiptDetail.objects.filter(
                receipt__student__standard=standard,
                receipt__student__academic_year=AcademicYear.objects.filter(is_current=True).first()
            ).select_related(
                'receipt', 
                'fee_type', 
                'fee_type__fee_master'
            ).values(
                'receipt__student_id',
                'fee_type__fee_master__name',
                'amount_paid',
                'amount_waived'
            )

            # Create a dictionary to store receipt totals by student and fee type
            receipt_totals = {}
            for detail in receipt_details:
                student_id = detail['receipt__student_id']
                fee_type_name = detail['fee_type__fee_master__name']
                
                if student_id not in receipt_totals:
                    receipt_totals[student_id] = {}
                
                if fee_type_name not in receipt_totals[student_id]:
                    receipt_totals[student_id][fee_type_name] = {
                        'paid': 0,
                        'waived': 0
                    }
                
                receipt_totals[student_id][fee_type_name]['paid'] += detail['amount_paid']
                receipt_totals[student_id][fee_type_name]['waived'] += detail['amount_waived']

            # Process student fees and receipts
            student_data = {}
            for fee in student_fees_queryset:
                student_id = fee.student_id
                if student_id not in student_data:
                    student_data[student_id] = {
                        'gr_no': fee.student.grno,
                        'first_name': fee.student.first_name,
                        'last_name': fee.student.last_name,
                        'standard': fee.standard.name if fee.standard else '',
                        'total_fee': 0,
                        'paid': 0,
                        'waived': 0,
                        'fee_structures': {}
                    }
                
                fee_type_name = str(fee.fee_type.fee_master) if fee.fee_type and fee.fee_type.fee_master else f'Fee Type {fee.fee_type_id}'
                fee_amount = fee.fee_type.amount if fee.fee_type else 0
                
                # Get paid and waived amounts from receipts
                paid_amount = 0
                waived_amount = 0
                if student_id in receipt_totals and fee_type_name in receipt_totals[student_id]:
                    paid_amount = receipt_totals[student_id][fee_type_name]['paid']
                    waived_amount = receipt_totals[student_id][fee_type_name]['waived']

                if fee_type_name not in student_data[student_id]['fee_structures']:
                    student_data[student_id]['fee_structures'][fee_type_name] = {
                        'amount': fee_amount,
                        'paid': paid_amount,
                        'waived': waived_amount,
                        'pending': fee_amount - paid_amount - waived_amount
                    }
                
                student_data[student_id]['total_fee'] += fee_amount
                student_data[student_id]['paid'] += paid_amount
                student_data[student_id]['waived'] += waived_amount

            # Calculate pending amount and prepare final list
            student_fees_list = []
            for student in student_data.values():
                student['pending'] = student['total_fee'] - student['paid'] - student['waived']
                student_fees_list.append(student)

            # Calculate totals
            totals = {
                'total_fee': sum(s['total_fee'] for s in student_fees_list),
                'paid': sum(s['paid'] for s in student_fees_list),
                'waived': sum(s['waived'] for s in student_fees_list),
                'pending': sum(s['pending'] for s in student_fees_list),
            }

            # Get all unique fee types
            fee_types = sorted([
                {
                    'id': fee.fee_type.fee_master.id,
                    'name': str(fee.fee_type.fee_master)
                }
                for fee in student_fees_queryset
                if fee.fee_type and fee.fee_type.fee_master
            ], key=lambda x: x['name'])

            # Remove duplicates while preserving order
            seen = set()
            fee_types = [x for x in fee_types if x['id'] not in seen and not seen.add(x['id'])]

            response_data = {
                'message': 'Fee report details retrieved successfully',
                'data': {
                    'students': student_fees_list,
                    'fee_types': fee_types,
                    'totals': totals,
                    'standard': standard
                }
            }

            return JsonResponse(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return JsonResponse(
                {'message': 'An error occurred', 'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
            
class FeeReportExcelView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasFeeReportPermission]
    required_permission = 'can_view_fee_report_detail'

    def get(self, request, standard):
        try:
            # Get student fees data with fee type details
            student_fees_queryset = student_fees.objects.filter(
                standard__name=standard,
                is_assigned=True,
                student__academic_year=AcademicYear.objects.filter(is_current=True).first()
            ).select_related(
                'student', 
                'standard', 
                'fee_type', 
                'fee_type__fee_master'
            )

            # Get receipt details
            receipt_details = ReceiptDetail.objects.filter(
                receipt__student__standard=standard,
                receipt__student__academic_year=AcademicYear.objects.filter(is_current=True).first()
            ).select_related(
                'receipt', 
                'fee_type', 
                'fee_type__fee_master'
            ).values(
                'receipt__student_id',
                'fee_type__fee_master__name',
                'amount_paid',
                'amount_waived'
            )

            # Create a dictionary to store receipt totals
            receipt_totals = {}
            for detail in receipt_details:
                student_id = detail['receipt__student_id']
                fee_type_name = detail['fee_type__fee_master__name']
                
                if student_id not in receipt_totals:
                    receipt_totals[student_id] = {}
                
                if fee_type_name not in receipt_totals[student_id]:
                    receipt_totals[student_id][fee_type_name] = {
                        'paid': 0,
                        'waived': 0
                    }
                
                receipt_totals[student_id][fee_type_name]['paid'] += detail['amount_paid']
                receipt_totals[student_id][fee_type_name]['waived'] += detail['amount_waived']

            # Process student fees and receipts
            student_data = {}
            for fee in student_fees_queryset:
                student_id = fee.student_id
                if student_id not in student_data:
                    student_data[student_id] = {
                        'gr_no': fee.student.grno,
                        'first_name': fee.student.first_name,
                        'last_name': fee.student.last_name,
                        'standard': fee.standard.name if fee.standard else '',
                        'total_fee': 0,
                        'paid': 0,
                        'waived': 0,
                        'fee_structures': {}
                    }
                
                fee_type_name = str(fee.fee_type.fee_master) if fee.fee_type and fee.fee_type.fee_master else f'Fee Type {fee.fee_type_id}'
                fee_amount = fee.fee_type.amount if fee.fee_type else 0
                
                # Get paid and waived amounts from receipts
                paid_amount = 0
                waived_amount = 0
                if student_id in receipt_totals and fee_type_name in receipt_totals[student_id]:
                    paid_amount = receipt_totals[student_id][fee_type_name]['paid']
                    waived_amount = receipt_totals[student_id][fee_type_name]['waived']

                if fee_type_name not in student_data[student_id]['fee_structures']:
                    student_data[student_id]['fee_structures'][fee_type_name] = {
                        'amount': fee_amount,
                        'paid': paid_amount,
                        'waived': waived_amount,
                        'pending': fee_amount - paid_amount - waived_amount
                    }
                
                student_data[student_id]['total_fee'] += fee_amount
                student_data[student_id]['paid'] += paid_amount
                student_data[student_id]['waived'] += waived_amount

            # Prepare final list and calculate totals
            student_fees_list = []
            for student in student_data.values():
                student['pending'] = student['total_fee'] - student['paid'] - student['waived']
                student_fees_list.append(student)

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            if standard == 13:
                ws.title = f"Fee Report - Balvatika"
                # Add title
                title = f"Fee Report - Standard Balvatika"
            else:
                ws.title = f"Fee Report - {standard}"
                # Add title
                title = f"Fee Report - Standard {standard}"
            ws.merge_cells('A1:F1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Get all unique fee types
            fee_types = sorted(list(set(
                fee_type
                for student in student_fees_list
                for fee_type in student['fee_structures'].keys()
            )))

            # Write headers
            headers = ['GR No', 'Name', 'Total Fee', 'Paid', 'Pending', 'Waived'] + fee_types
            ws.append([''] * len(headers))  # Add empty row for spacing
            ws.append(headers)
            ws.append(['', '', '', '', '', ''] + ['Total/Paid' for _ in fee_types])

            # Write student data
            for student in student_fees_list:
                row = [
                    student['gr_no'],
                    f"{student['first_name']} {student['last_name']}",
                    student['total_fee'],
                    student['paid'],
                    student['pending'],
                    student['waived']
                ]
                for fee_type in fee_types:
                    if fee_type in student['fee_structures']:
                        amount = student['fee_structures'][fee_type]['amount']
                        paid = student['fee_structures'][fee_type]['paid']
                        row.append(f"{amount}/{paid}")
                    else:
                        row.append("-")
                ws.append(row)

            # Calculate and write totals
            totals = {
                'total_fee': sum(s['total_fee'] for s in student_fees_list),
                'paid': sum(s['paid'] for s in student_fees_list),
                'pending': sum(s['pending'] for s in student_fees_list),
                'waived': sum(s['waived'] for s in student_fees_list),
            }

            totals_row = [
                'Total', '',
                totals['total_fee'],
                totals['paid'],
                totals['pending'],
                totals['waived']
            ]
            
            for fee_type in fee_types:
                total_amount = sum(s['fee_structures'].get(fee_type, {}).get('amount', 0) for s in student_fees_list)
                total_paid = sum(s['fee_structures'].get(fee_type, {}).get('paid', 0) for s in student_fees_list)
                totals_row.append(f"{total_amount}/{total_paid}")
            ws.append(totals_row)

            # Style the worksheet
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row[0].row > 1:  # Skip title row
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

            # Make headers bold
            for cell in ws[3] + ws[4]:  # Headers are in rows 3 and 4
                cell.font = Font(bold=True)

            # Adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells if not isinstance(cell, MergedCell))
                if length > 0:
                    column_letter = get_column_letter(column_cells[0].column)
                    ws.column_dimensions[column_letter].width = length + 2

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="fee_report_standard_{standard}.xlsx"'
            
            wb.save(response)
            return response

        except Exception as e:
            return JsonResponse(
                {'message': 'An error occurred', 'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )            
            
class FeeTypeReportExcelViewSingle(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, HasFeeReportPermission]
    required_permission = 'can_view_fee_report_detail'

    def get(self, request, standard, fee_master_id):
        try:
            # Create workbook and sheet
            wb = Workbook()
            ws = wb.active
            
            # Get fee type master name for title
            fee_master_obj = fee_type_master.objects.filter(id=fee_master_id).first()
            if not fee_master_obj:
                return JsonResponse(
                    {'message': 'Fee type master not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            fee_type_name = str(fee_master_obj.name)
            
            if standard.upper() == 'ALL':
                ws.title = f"{fee_type_name} - All Standards"
                title = f"Fee Report - All Standards - {fee_type_name}"
                # Get student fees for all standards
                student_fees_queryset = student_fees.objects.filter(
                    is_assigned=True,
                    fee_type__fee_master_id=fee_master_id,
                    student__academic_year=AcademicYear.objects.filter(is_current=True).first()
                ).select_related(
                    'student', 
                    'standard', 
                    'fee_type',
                    'fee_type__fee_master'
                )
            else:
                if standard == "13":
                    ws.title = f"{fee_type_name} - Balvatika"
                    title = f"Fee Report - Standard Balvatika - {fee_type_name}"
                else:
                    ws.title = f"{fee_type_name} - {standard}"
                    title = f"Fee Report - Standard {standard} - {fee_type_name}"
                # Get student fees for specific standard
                student_fees_queryset = student_fees.objects.filter(
                    standard__name=standard,
                    is_assigned=True,
                    fee_type__fee_master_id=fee_master_id,
                    student__academic_year=AcademicYear.objects.filter(is_current=True).first()
                ).select_related(
                    'student', 
                    'standard', 
                    'fee_type',
                    'fee_type__fee_master'
                )

            # Add title
            ws.merge_cells('A1:G1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Add an empty row for spacing
            ws.append([''])

            # Add headers
            headers = ['GR No', 'Name', 'Standard', 'Amount', 'Paid', 'Waived', 'Pending']
            ws.append(headers)

            # Get receipt details
            if standard.upper() == 'ALL':
                receipt_details = ReceiptDetail.objects.filter(
                    fee_type__fee_master_id=fee_master_id,
                    receipt__student__academic_year=AcademicYear.objects.filter(is_current=True).first()
                )
            else:
                receipt_details = ReceiptDetail.objects.filter(
                    receipt__student__standard=standard,
                    fee_type__fee_master_id=fee_master_id,
                    receipt__student__academic_year=AcademicYear.objects.filter(is_current=True).first()
                )

            receipt_details = receipt_details.values(
                'receipt__student_id',
                'amount_paid',
                'amount_waived'
            )

            # Create a dictionary to store receipt totals
            receipt_totals = {}
            for detail in receipt_details:
                student_id = detail['receipt__student_id']
                if student_id not in receipt_totals:
                    receipt_totals[student_id] = {
                        'paid': 0,
                        'waived': 0
                    }
                receipt_totals[student_id]['paid'] += detail['amount_paid']
                receipt_totals[student_id]['waived'] += detail['amount_waived']

            # Process student fees and add to Excel
            total_amount = total_paid = total_waived = total_pending = 0

            for fee in student_fees_queryset:
                student_id = fee.student_id
                amount = fee.fee_type.amount if fee.fee_type else 0
                paid = receipt_totals.get(student_id, {}).get('paid', 0)
                waived = receipt_totals.get(student_id, {}).get('waived', 0)
                pending = amount - paid - waived

                ws.append([
                    fee.student.grno,
                    f"{fee.student.first_name} {fee.student.last_name}",
                    fee.standard.name if fee.standard else '',
                    amount,
                    paid,
                    waived,
                    pending
                ])

                total_amount += amount
                total_paid += paid
                total_waived += waived
                total_pending += pending

            # Add totals row
            ws.append([
                'Total', '', '',
                total_amount,
                total_paid,
                total_waived,
                total_pending
            ])

            # Apply styles
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row[0].row > 2:  # Skip applying border to the title and empty rows
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

            # Make headers bold
            for cell in ws[3]:  # Headers are in row 3
                cell.font = Font(bold=True)

            # Adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells if not isinstance(cell, MergedCell))
                if length > 0:
                    column_letter = get_column_letter(column_cells[0].column)
                    ws.column_dimensions[column_letter].width = length + 2

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="fee_report_{standard}_{fee_type_name}.xlsx"'
            
            wb.save(response)
            return response

        except Exception as e:
            return JsonResponse(
                {'message': 'An error occurred', 'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# email

class PermissionRequestView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get request_id from query parameters
            request_id = request.query_params.get('request_id')
            
            if request_id:
                # Get specific request status
                cache_key = f'permission_request_{request_id}'
                request_data = cache.get(cache_key)
                
                if not request_data:
                    return Response({
                        'message': 'Permission request not found or expired',
                        'status': 'expired',
                        'request_id': request_id
                    }, status=status.HTTP_404_NOT_FOUND)
                
                try:
                    group = Group.objects.get(name=request_data['group_name'])
                    user = User.objects.get(id=request_data['user_id'])
                    
                    # Check if user has the permission
                    has_permission = group in user.groups.all()
                    current_status = 'approved' if has_permission else request_data['status']
                    
                    return Response({
                        'message': 'Permission request status retrieved',
                        'details': {
                            'request_id': request_id,
                            'status': current_status,
                            'user': user.email,
                            'group': group.name,
                            'reason': request_data.get('reason', ''),
                            'submitted_at': request_data.get('submitted_at', ''),
                            'admin_emails': request_data.get('admin_emails', []),
                            'has_permission': has_permission
                        }
                    }, status=status.HTTP_200_OK)
                    
                except (Group.DoesNotExist, User.DoesNotExist):
                    return Response({
                        'message': 'Associated group or user not found',
                        'status': request_data['status']
                    }, status=status.HTTP_200_OK)

            # Get available groups and admin users
            groups = Group.objects.all().values('name', 'permissions__name').distinct()
            admin_users = User.objects.filter(is_staff=True).values('id', 'email', 'username')

            # Format groups data
            formatted_groups = {}
            for group in groups:
                group_name = group['name']
                if group_name not in formatted_groups:
                    formatted_groups[group_name] = {
                        'name': group_name,
                        'permissions': []
                    }
                if group['permissions__name']:
                    formatted_groups[group_name]['permissions'].append(group['permissions__name'])

            return Response({
                'groups': list(formatted_groups.values()),
                'admin_users': admin_users
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request):
        try:
            group_name = request.data.get('group_name')
            admin_ids = request.data.get('admin_ids', [])
            reason = request.data.get('reason', '')

            if not group_name:
                return Response(
                    {'error': 'Group name is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                group = Group.objects.get(name=group_name)
            except Group.DoesNotExist:
                return Response(
                    {'error': 'Group does not exist'}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Check if user already has this group
            if group in request.user.groups.all():
                return Response({
                    'message': 'You already have this group permission',
                    'details': {
                        'user': request.user.email,
                        'group': group.name,
                        'status': 'approved',
                        'has_permission': True,
                        'permissions': list(group.permissions.values_list('name', flat=True))
                    }
                }, status=status.HTTP_200_OK)

            # Get admin emails
            admin_emails = []
            if admin_ids:
                admin_emails = User.objects.filter(
                    id__in=admin_ids,
                    is_staff=True
                ).values_list('email', flat=True)
            else:
                admin_emails = User.objects.filter(
                    is_staff=True
                ).values_list('email', flat=True)

            if not admin_emails:
                return Response(
                    {'error': 'No admin users found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Generate unique request ID
            request_id = str(uuid.uuid4())
            
            # Store request data in cache
            cache_data = {
                'user_id': request.user.id,
                'group_name': group.name,
                'status': 'pending',
                'reason': reason,
                'admin_emails': list(admin_emails),
                'submitted_at': datetime.datetime.now().isoformat()
            }
            
            # Store the request data with group name
            cache.set(f'permission_request_{request_id}', cache_data, 60 * 60 * 48)
            cache.set(f'user_requests_{request.user.id}_{group.name}', request_id, 60 * 60 * 48)

            # Generate approval/decline URLs
            base_url = request.build_absolute_uri('/')[:-1]
            approve_url = f"{base_url}/permission-request/approve/{request_id}/"
            decline_url = f"{base_url}/permission-request/decline/{request_id}/"

            # Send email
            subject = f'Permission Request from {request.user.email}'
            message = f"""
            User Details:
            - Email: {request.user.email}
            - Username: {request.user.username}
            
            Has requested access to group: {group.name}
            
            Group Permissions:
            {', '.join(group.permissions.values_list('name', flat=True))}
            
            Reason for request: {reason}
            
            Actions:
            - Approve: {approve_url}
            - Decline: {decline_url}
            
            This request will expire in 48 hours.
            """
            
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                list(admin_emails),
                fail_silently=False,
            )

            return Response({
                'message': 'Permission request sent successfully',
                'details': {
                    'request_id': request_id,
                    'group': group.name,
                    'admin_emails': list(admin_emails),
                    'status': 'pending'
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

            
            
class ApprovePermissionView(APIView):
    def get(self, request, request_id):
        try:
            # Get request data from cache
            cache_key = f'permission_request_{request_id}'
            request_data = cache.get(cache_key)

            if not request_data:
                return Response(
                    {'error': 'Permission request not found or expired'}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            if request_data['status'] != 'pending':
                return Response(
                    {'error': 'This request has already been processed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                # Get user and group using name instead of id
                user = User.objects.get(id=request_data['user_id'])
                group = Group.objects.get(name=request_data['group_name'])

                # Check if user already has this group
                if group in user.groups.all():
                    return Response({
                        'message': 'User already has this permission',
                        'details': {
                            'user': user.email,
                            'group': group.name,
                            'status': 'approved'
                        }
                    }, status=status.HTTP_200_OK)

                # Add user to group
                user.groups.add(group)
                
                # Update cache status
                request_data['status'] = 'approved'
                cache.set(cache_key, request_data, 60 * 60 * 48)

                # Send confirmation email to user and all admins
                subject = 'Permission Request Approved'
                user_message = f"""
                Your request for access to group '{group.name}' has been approved.
                
                You now have access to the following permissions:
                {', '.join(group.permissions.values_list('name', flat=True))}
                """
                
                admin_message = f"""
                The permission request from {user.email} for group '{group.name}' has been approved.
                
                Request Details:
                - User: {user.email}
                - Group: {group.name}
                - Reason: {request_data.get('reason', 'No reason provided')}
                - Permissions Granted: {', '.join(group.permissions.values_list('name', flat=True))}
                """
                
                # Send to user
                send_mail(
                    subject,
                    user_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email],
                    fail_silently=False,
                )

                # Send to admins
                send_mail(
                    subject,
                    admin_message,
                    settings.DEFAULT_FROM_EMAIL,
                    request_data['admin_emails'],
                    fail_silently=False,
                )

                return Response({
                    'message': 'Permission request approved successfully',
                    'details': {
                        'user': user.email,
                        'group': group.name,
                        'status': 'approved',
                        'permissions': list(group.permissions.values_list('name', flat=True))
                    }
                }, status=status.HTTP_200_OK)

            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            except Group.DoesNotExist:
                return Response(
                    {'error': 'Group not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )

        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DeclinePermissionView(APIView):
    def get(self, request, request_id):
        try:
            # Get request data from cache
            cache_key = f'permission_request_{request_id}'
            request_data = cache.get(cache_key)

            if not request_data:
                return Response(
                    {'error': 'Permission request not found or expired'}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            if request_data['status'] != 'pending':
                return Response({
                    'message': 'This request has already been processed',
                    'details': {
                        'status': request_data['status'],
                        'note': 'Permission was already handled'
                    }
                }, status=status.HTTP_200_OK)

            # Get user and group
            try:
                user = User.objects.get(id=request_data['user_id'])
                group = Group.objects.get(id=request_data['group_id'])
                
                # Check if user already has this group
                if group in user.groups.all():
                    return Response({
                        'message': 'Group permission already exists',
                        'details': {
                            'user': user.email,
                            'group': group.name,
                            'status': 'already_exists',
                            'note': 'User already has this group permission'
                        }
                    }, status=status.HTTP_200_OK)

                # Assign the group to the user
                user.groups.add(group)
                user.save()

                # Verify the group was assigned
                if group not in user.groups.all():
                    raise Exception("Failed to assign group to user")

                # Update cache status
                request_data['status'] = 'approved'
                cache.set(cache_key, request_data, 60 * 60 * 48)

                # Send confirmation emails
                subject = 'Permission Request Approved'
                user_message = f"""
                Your request for access to group '{group.name}' has been approved.
                
                You now have access to the requested permissions.
                Group Details:
                - Group Name: {group.name}
                - Permissions Granted: {', '.join(group.permissions.values_list('name', flat=True))}
                """
                
                admin_message = f"""
                New permission has been granted successfully.
                
                Details:
                - User: {user.email} ({user.username})
                - Group: {group.name}
                - Reason: {request_data.get('reason', 'No reason provided')}
                - Status: Successfully assigned
                """
                
                # Send emails
                send_mail(
                    subject,
                    user_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email],
                    fail_silently=False,
                )

                send_mail(
                    subject,
                    admin_message,
                    settings.DEFAULT_FROM_EMAIL,
                    request_data['admin_emails'],
                    fail_silently=False,
                )

                return Response({
                    'message': 'New group permission assigned successfully',
                    'details': {
                        'user': user.email,
                        'group': group.name,
                        'status': 'newly_approved',
                        'permissions_granted': list(group.permissions.values_list('name', flat=True))
                    }
                }, status=status.HTTP_200_OK)

            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            except Group.DoesNotExist:
                return Response(
                    {'error': 'Group not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            except Exception as e:
                return Response(
                    {'error': f'Failed to assign group: {str(e)}'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )